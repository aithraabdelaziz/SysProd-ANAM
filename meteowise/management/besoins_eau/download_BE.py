import os
import numpy as np
import pandas as pd
import geopandas as gpd
import matplotlib.pyplot as plt
from pykrige.ok import OrdinaryKriging
from shapely.geometry import Point
from matplotlib.colors import ListedColormap
from cartopy import crs as ccrs
from cartopy.feature import ShapelyFeature
from cartopy.io.shapereader import Reader
from sklearn.metrics import mean_squared_error
from scipy.optimize import curve_fit
from scipy.spatial.distance import pdist, squareform
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from sklearn.metrics import mean_squared_error
import os
import pandas as pd
from scipy.spatial.distance import pdist, squareform
import geopandas as gpd
import itertools
import openpyxl
import csv
import openmeteo_requests
import requests_cache
from retry_requests import retry
import xarray as xr
import rasterio
from rasterio.transform import from_origin
from retry_requests import retry

from pprint import pprint
from datetime import date, datetime
import unicodedata

def remove_accents(text):
    nfkd_form = unicodedata.normalize('NFKD', text)
    # Supprime les marques diacritiques (accents)
    return ''.join([c for c in nfkd_form if not unicodedata.combining(c)])
def get_current_decade_code(date_now=None):
    if date_now is not None : today = date_now.date()
    else : today = date.today()
    day = today.day
    month = today.month
    year = today.year

    # Calcul de la décade (01 = 1-10, 02 = 11-20, 03 = 21-fin)
    if day <= 10:
        decade = "01"
    elif day <= 20:
        decade = "02"
    else:
        decade = "03"

    # Format final : JJMMYYYY
    code = f"{decade}{month:02d}{year}"
    return code

# Configuration du client Open-Meteo avec mise en cache et retry en cas d'erreur
cache_session = requests_cache.CachedSession('.cache', expire_after=3600)
retry_session = retry(cache_session, retries=5, backoff_factor=0.2)
openmeteo = openmeteo_requests.Client(session=retry_session)

# URL de l'API Open-Meteo
url = "https://api.open-meteo.com/v1/forecast"

# Dictionnaire des 10 points avec noms et coordonnées (ordre important)
locations = {
    "BOBO": (11.17, -4.3),
    "BOGANDE": (12.98, -0.13),
    "BOROMO": (11.73, -2.92),
    "DEDOUGOU": (12.47, -3.48),
    "DORI": (14.03, -0.03),
    "FADA": (12.07, 0.35),
    "GAOUA": (10.33, -3.18),
    "OUAGA": (12.35, -1.52),
    "OUAHIGOUYA": (13.58, -2.43),
    "PO": (11.17, -1.15),
    "BEREGADOUGOU": (10.75, -4.79)
}

# Initialiser une liste pour stocker les DataFrames
all_dataframes = []
current_decade=  get_current_decade_code()

now = datetime.now()
print(now.strftime("[%Y-%m-%d %H:%M:%S]")+f" INFO Calcul des besoins en eau pour la décade {current_decade}")
# results_dir = f"results/{current_decade}"
pathdata = os.path.dirname(os.path.abspath(__file__))
results_dir = f"{pathdata}/../../../media/agromet/besoins_eau/{current_decade}"
os.makedirs(results_dir, exist_ok=True)
constantes_dir = f"{pathdata}/constantes"

# Boucle sur chaque emplacement avec index pour conserver l'ordre
for order, (location_name, (lat, lon)) in enumerate(locations.items()):
    params = {
        "latitude": lat,
        "longitude": lon,
        "daily": "et0_fao_evapotranspiration",
        "forecast_days": 10,
        "models": "gfs_seamless"
    }
    
    # Appel de l'API Open-Meteo
    responses = openmeteo.weather_api(url, params=params)
    
    # Récupération de la réponse pour ce point
    response = responses[0]
    
    # Extraction des données journalières
    daily = response.Daily()
   
    daily_et0_fao_evapotranspiration = daily.Variables(0).ValuesAsNumpy()
    
    # Création d'un DataFrame
    daily_data = {
        "date": pd.date_range(
            start=pd.to_datetime(daily.Time(), unit="s", utc=True),
            end=pd.to_datetime(daily.TimeEnd(), unit="s", utc=True),
            freq=pd.Timedelta(seconds=daily.Interval()),
            inclusive="left"
        ),
        "location_name": location_name,  # Ajout du nom du point
        "latitude": lat,
        "longitude": lon,
        "et0_fao_evapotranspiration": daily_et0_fao_evapotranspiration,
        "order": order  # Ajout de l'ordre pour tri final
    }
    df = pd.DataFrame(daily_data)
    
    # Ajouter la période décadaire (premier jour de la période)
    df["decade"] = df["date"].min().strftime("%Y-%m-%d")

    # Conserver l'ordre établi
    all_dataframes.append(df)

# Concaténer toutes les données en un seul DataFrame
final_dataframe = pd.concat(all_dataframes, ignore_index=True)

# Calcul du cumul décadaire par emplacement et période
decadal_cumsum = (
    final_dataframe.groupby(["location_name", "latitude", "longitude", "decade", "order"])["et0_fao_evapotranspiration"]
    .sum()
    .reset_index()
)

# Trier les résultats selon l'ordre initial
decadal_cumsum = decadal_cumsum.sort_values(by="order").drop(columns=["order"])  # Suppression de la colonne après tri

# Sauvegarde dans un fichier CSV
csv_file = f"{results_dir}/et0_forecast_decadal.csv"

# Écriture du fichier CSV avec encodage UTF-8 et séparation par ";"
decadal_cumsum.to_csv(csv_file, index=False, sep=";", encoding="utf-8")

# print(f"Données enregistrées dans {csv_file} avec l'ordre respecté.")
# Fichiers source
# csv_file = "./bad_etp_PREV/et0_forecast_decadal_01-08_03_2025.csv"
excel_file = f"{constantes_dir}/et0_par_stations.xltx"

# --- Lecture du CSV avec le bon séparateur (virgule) ---
with open(csv_file, 'r', newline='', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f, delimiter=';')  # <-- Séparateur modifié de ';' à ','
    csv_data = list(reader)

# Debug : afficher les clés du CSV pour vérifier le header
if not csv_data:
    raise ValueError("Le fichier CSV est vide.")
# else:
#     print("Clés du CSV :", list(csv_data[0].keys()))

# Vérifier que la colonne 'et0_fao_evapotranspiration' existe bien
if 'et0_fao_evapotranspiration' not in csv_data[0]:
    raise ValueError("La colonne 'et0_fao_evapotranspiration' n'existe pas dans le CSV.")

# Extraire les valeurs de la colonne
et0_values = [row['et0_fao_evapotranspiration'] for row in csv_data]

# --- Traitement du fichier Excel avec openpyxl ---
# Charger le fichier Excel (.xltx)
wb = openpyxl.load_workbook(excel_file)
ws = wb.active  # On prend le premier onglet

# Recherche de la colonne 'et0' dans la première ligne (les en-têtes)
et0_column_index = None
for cell in ws[1]:
    if cell.value is not None and str(cell.value).strip() == 'et0':
        et0_column_index = cell.column  # openpyxl retourne un entier pour la colonne
        break

# Si la colonne 'et0' n'est pas trouvée, on la crée en position A (colonne 1)
if et0_column_index is None:
    print("La colonne 'et0' n'a pas été trouvée dans le fichier Excel. Création en colonne A.")
    et0_column_index = 1
    ws.cell(row=1, column=et0_column_index, value='et0')

# Copier les valeurs dans la colonne 'et0' à partir de la ligne 2
for i, value in enumerate(et0_values, start=2):
    ws.cell(row=i, column=et0_column_index, value=value)

# Enregistrer le fichier Excel mis à jour
wb.save(excel_file)
# print(f"Mise à jour enregistrée dans '{excel_file}'.")

# Chemins des fichiers
et0_file_path = excel_file #'et0_par_stations.xltx'
kc_files = {
    'Maïs': f'{constantes_dir}/kc_par_decade_mais.xltx',
    'Tomate': f'{constantes_dir}/kc_par_decade_tomate.xltx',
    'Oignon': f'{constantes_dir}/kc_par_decade_oignon.xltx',
    'Choux': f'{constantes_dir}/kc_par_decade_choux.xltx'
}

# Chargement des données d'ET0
et0_data = pd.read_excel(et0_file_path, sheet_name='Feuil1')
stations = et0_data['stations']
et0_values = et0_data['et0']

# Fonction pour trier les décades selon l'ordre naturel
def sort_decades(decade_list):
    return sorted(decade_list, key=lambda x: int(x.split('-')[0]))

# Fonction pour traiter les besoins en eau pour une culture donnée
def calculate_water_requirements(culture_name, kc_file_path):
    # Charger les coefficients Kc
    kc_data = pd.read_excel(kc_file_path, sheet_name='Feuil1')
    kc_data['decade'] = kc_data['decade'].astype(str)  # Transformation explicite en chaînes
    decades = kc_data['decade']
    kc_values = kc_data['kc']

    # Création du produit cartésien pour combiner les stations avec les décades
    combinations = pd.DataFrame(list(itertools.product(stations, decades)), columns=['Station', 'Décade'])

    # Associer les valeurs ET0 et Kc
    combinations['ET0 (mm)'] = combinations['Station'].map(dict(zip(stations, et0_values)))
    combinations['Kc'] = combinations['Décade'].map(dict(zip(decades, kc_values)))

    # Calcul des besoins en eau (ETc)
    combinations['Besoins en eau (mm)'] = combinations['ET0 (mm)'] * combinations['Kc']

    # Conversion des besoins en eau en nombre d'arrosoirs arrondi à l'entier
    combinations['Arrosoirs (pour 25 m²)'] = (combinations['Besoins en eau (mm)'] * 25 / 150).round().astype(int)

    # Réorganiser le tableau avec les stations en lignes et les décades en colonnes
    arrosoirs_pivot = combinations.pivot(index='Station', columns='Décade', values='Arrosoirs (pour 25 m²)')

    # Tri explicite des colonnes (décades) dans l'ordre naturel
    arrosoirs_pivot = arrosoirs_pivot.reindex(sort_decades(arrosoirs_pivot.columns), axis=1)
    
    return arrosoirs_pivot
# Fonction pour exporter un tableau sous forme d'image PNG
def export_to_png(dataframe, culture_name, output_dir=f'{results_dir}/png'):
    os.makedirs(output_dir, exist_ok=True)
    # Préparer les colonnes
    col_labels = ['Provinces'] + [f' {decade}' for decade in dataframe.columns]

    # Exportation en PNG avec une police augmentée pour les nombres
    fig, ax = plt.subplots(figsize=(5, 3.4))
    ax.axis('tight')
    ax.axis('off')

    # Ajouter un titre au tableau
    fig.suptitle(
    f"Besoins en eau (nombre arrosoirs/jour pour 25 m²) pour {culture_name}.\n\n{' ' * 115}Différentes phases de la culture exprimées en nombre de jours après semis sur la première ligne",
    fontsize=11,
    fontweight='bold',
    y=0.95
    )

    # Création du tableau avec une police augmentée
    table = ax.table(cellText=dataframe.reset_index().values, colLabels=col_labels, cellLoc='center', loc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(12)  # Taille de police ajustée
    table.auto_set_column_width(col=list(range(len(col_labels))))

    # Sauvegarde du tableau sous format PNG
    culture_name=remove_accents(culture_name)
    output_path = os.path.join(output_dir, f'{culture_name.lower()}.png')
    plt.savefig(output_path, bbox_inches='tight', dpi=500)
    # print(f"Tableau pour {culture_name} exporté sous format PNG à l'emplacement : {output_path}")
    plt.close(fig)

def export_to_html(dataframe, culture_name, output_dir=f'{results_dir}/html'):
    os.makedirs(output_dir, exist_ok=True)

    # Préparer les colonnes
    col_labels = ['Provinces'] + [f' {decade}' for decade in dataframe.columns]
    df_html = dataframe.reset_index()
    df_html.columns = col_labels

    # Construire le titre
    title = (
        f"<center>"
        f"<strong>Besoins en eau (nombre arrosoirs/jour pour 25 m²) pour {culture_name}.</strong>"
        f"<br><small>Différentes phases de la culture exprimées en nombre de jours après semis sur la première ligne</small>"
        f"</center>"
    )

    # Générer la table HTML
    table_html = df_html.to_html(index=False, border=0, classes='water-needs-table', justify='center')

    # Mettre en forme avec un peu de style CSS intégré
    full_html = f"""
    <div>
        <style>

            .water-needs-table {{
                border-collapse: collapse;
                width: 100%;
                table-layout: auto; 
                max-width: 100%; 
                border-collapse: collapse;
                font-size:8px;
            }}
            .water-needs-table th, .water-needs-table td {{
                border: 1px solid #ddd;
                padding: 0px;
                text-align: center;
            }}
            .water-needs-table th {{
                background-color: #f2f2f2;
            }}
        </style>
    
        {title}
        {table_html}
    </div>
    """

    # full_html = f"{title}<br>{table_html}"

    # Sauvegarde
    culture_name=remove_accents(culture_name)
    output_path = os.path.join(output_dir, f'{culture_name.lower()}.html')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(full_html)

    # print(f"Tableau HTML pour {culture_name} exporté à : {output_path}")

# Calcul et exportation pour chaque culture
for culture_name, kc_file_path in kc_files.items():
    arrosoirs_pivot = calculate_water_requirements(culture_name, kc_file_path)
    export_to_png(arrosoirs_pivot, culture_name)
    export_to_html(arrosoirs_pivot, culture_name)